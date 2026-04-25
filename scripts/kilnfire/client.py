"""Thin Kilnfire admin client. Lifted from the MPS sub-repo's
kilnfire_fetch.py — only the bits needed for the daily scrape.

Only READS Kilnfire — never writes. Per house rule: writes against
Kilnfire happen only via the UI by Nick, and only against Sept 2026
"TEST —" classes.
"""

import os, requests
from bs4 import BeautifulSoup


def login_session():
    """Return a requests.Session authenticated to Kilnfire.

    Reads KILNFIRE_URL / KILNFIRE_USER / KILNFIRE_PASS from env.
    """
    base = os.environ["KILNFIRE_URL"].rstrip("/")
    s = requests.Session()
    s.headers.update({"User-Agent": "Mozilla/5.0"})

    r = s.get(f"{base}/login")
    soup = BeautifulSoup(r.text, "html.parser")
    token = soup.find("input", {"name": "_token"})["value"]

    s.post(
        f"{base}/login",
        data={
            "_token": token,
            "email": os.environ["KILNFIRE_USER"],
            "password": os.environ["KILNFIRE_PASS"],
        },
        allow_redirects=True,
    )
    return s, token, base


def fetch_planning(session, token, base):
    """Pull the planning/data endpoint — full list of class instances
    (past + future). Returns the parsed JSON payload."""
    headers = {
        "Accept": "application/json",
        "X-Requested-With": "XMLHttpRequest",
        "X-CSRF-TOKEN": token,
        "Content-Type": "application/json",
    }
    r = session.get(f"{base}/classes/admin/planning/data", headers=headers)
    r.raise_for_status()
    return r.json()


def fetch_planning_csv(session, token, base):
    """Alternate path — Kilnfire serves the same data as a CSV export.
    Useful when the data endpoint shape changes. Returns CSV text."""
    headers = {"Accept": "*/*", "X-CSRF-TOKEN": token}
    r = session.get(f"{base}/classes/admin/planning/export", headers=headers)
    r.raise_for_status()
    return r.text


def fetch_planning_rows(session, token, base):
    """Fetch the planning export and parse it into a list of dicts. Kilnfire
    actually returns XLSX from this endpoint despite the .csv path; detect
    the format by the response content-type and ZIP magic bytes."""
    from io import BytesIO
    from openpyxl import load_workbook

    headers = {"Accept": "*/*", "X-CSRF-TOKEN": token}
    r = session.get(f"{base}/classes/admin/planning/export", headers=headers)
    r.raise_for_status()

    body = r.content
    # ZIP/XLSX files start with 'PK\x03\x04'
    if body[:4] == b'PK\x03\x04':
        wb = load_workbook(BytesIO(body), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers_row = [str(h) if h is not None else '' for h in rows[0]]
        return [
            {headers_row[i]: (cell if cell is not None else '') for i, cell in enumerate(r)}
            for r in rows[1:]
        ]

    # Otherwise treat as CSV.
    import csv, io
    text = body.decode('utf-8', errors='replace').replace('\r\n', '\n').replace('\r', '\n')
    return list(csv.DictReader(io.StringIO(text)))
